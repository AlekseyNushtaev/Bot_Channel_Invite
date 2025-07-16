import asyncio

from aiogram import types, F, Router
from aiogram.filters import ChatMemberUpdatedFilter, KICKED, MEMBER, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import default_state, StatesGroup, State
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ChatMemberUpdated
from openpyxl import Workbook
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from openpyxl.utils import get_column_letter
from io import BytesIO


from config import ADMIN_IDS
from db.utils import add_subscription_request, confirm_subscription, update_user_unblocked, \
    update_user_blocked, get_all_users_unblock

from bot import bot
from db.models import SessionLocal, SubscriptionRequest
from keyboard import create_kb, kb_button

router = Router()


class FSMFillForm(StatesGroup):
    get_phone = State()
    send = State()
    text_add_button = State()
    check_text_1 = State()
    check_text_2 = State()
    text_add_button_text = State()
    text_add_button_url = State()
    photo_add_button = State()
    check_photo_1 = State()
    check_photo_2 = State()
    photo_add_button_text = State()
    photo_add_button_url = State()
    video_add_button = State()
    check_video_1 = State()
    check_video_2 = State()
    video_add_button_text = State()
    video_add_button_url = State()
    check_video_note_1 = State()
    send_id = State()
    send_to_one = State()


# Обработка заявки на вступление
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton

# Обработка заявки на вступление
@router.chat_join_request()
async def handle_join_request(join_request: types.ChatJoinRequest):
    user = join_request.from_user
    chat = join_request.chat
    # try:
    #     invite_link = await bot.export_chat_invite_link(chat_id=chat.id)  # Полная ссылка
    # except Exception:
    invite_link = None

    # Сохраняем заявку в БД
    add_subscription_request(
        user_id=user.id,
        username=user.username,
        first_name=user.first_name,
        last_name=user.last_name,
        channel_id=chat.id,
        channel_name=chat.title,
        channel_link=invite_link
    )

    # Создаем Reply кнопку
    subscribe_button = KeyboardButton(text="Подтвердить")
    reply_markup = ReplyKeyboardMarkup(
        keyboard=[[subscribe_button]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

    # Отправляем сообщение с кнопкой
    send_msg = await bot.send_message(
        chat_id=user.id,
        text=f"Подтвердите что вы не робот 🤖",
        reply_markup=reply_markup
    )
    await asyncio.sleep(60)
    try:
        # Удаляем сообщение с кнопкой "Подтвердить"
        await send_msg.delete()
    except Exception as e:
        pass


# Обработка нажатия кнопки "Подписаться"
@router.message(F.text == "Подтвердить")
async def handle_subscribe_button(message: types.Message):
    user = message.from_user

    # Ищем последнюю активную заявку
    db = SessionLocal()
    try:
        request = db.query(SubscriptionRequest).filter(
            SubscriptionRequest.user_id == user.id,
            SubscriptionRequest.time_confirm.is_(None),
        ).order_by(SubscriptionRequest.time_request.desc()).first()

        if not request:
            # await message.answer("❌ Нет активных заявок.")
            return

        # Одобряем заявку
        try:
            # await bot.approve_chat_join_request(
            #     chat_id=request.channel_id,
            #     user_id=user.id,
            # )
            confirm_subscription(user.id, request.channel_id)

            # Создаем ссылку на канал
            # channel_link = request.channel_link

            # Формируем сообщение с кнопкой
            # text = f"Ваша заявка на вступление на рассмотрение !"
            # reply_markup = None
            #
            # if channel_link:
            #     reply_markup = InlineKeyboardMarkup(inline_keyboard=[
            #         [InlineKeyboardButton(text="Перейти в канал", url=channel_link)]
            #     ])

            confirmation_msg = await message.answer(
                text="Ваша заявка на вступление на рассмотрении !",
                reply_markup=None
            )


            # Убираем Reply клавиатуру
            # await message.answer("Спасибо за подписку!", reply_markup=types.ReplyKeyboardRemove())
            # Удаляем исходное сообщение с кнопкой и предыдущие сообщения через минуту
            await asyncio.sleep(60)  # Ждем минуту

            try:
                # Удаляем сообщение с кнопкой "Подтвердить"
                await message.delete()
            except Exception as e:
                pass
                # Удаляем сообщение бота с кнопкой
            try:
                await bot.delete_message(chat_id=user.id, message_id=message.message_id - 1)
            except Exception as e:
                pass
                # Удаляем сообщение "Ваша заявка на рассмотрении"
            try:
                await confirmation_msg.delete()
            except Exception as e:
                pass
        except Exception as e:
            pass
            # await message.answer("❌ Ошибка. Возможно, заявка уже обработана или отменена.")
    finally:
        db.close()


@router.my_chat_member(ChatMemberUpdatedFilter(member_status_changed=KICKED))
async def user_blocked_bot(event: ChatMemberUpdated):
    update_user_blocked(event.from_user.id)


@router.my_chat_member(ChatMemberUpdatedFilter(member_status_changed=MEMBER))
async def user_unblocked_bot(event: ChatMemberUpdated):
    update_user_unblocked(event.from_user.id)



@router.message(F.text == 'Send', StateFilter(default_state), F.from_user.id.in_(ADMIN_IDS))
async def send_to_all(message: types.Message, state: FSMContext):
    await message.answer(text='Сейчас мы подготовим сообщение для рассылки по юзерам!\n'
                              'Отправьте пжл текстовое сообщение или картинку(можно с текстом) или видео(можно с текстом) или видео-кружок')
    await state.set_state(FSMFillForm.send)


@router.message(F.text, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button(message: types.Message, state: FSMContext):
    await state.update_data(text=message.text)
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.text_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.text_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    await cb.message.answer(text=dct['text'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_text_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_text_1), F.from_user.id.in_(ADMIN_IDS))
async def check_text_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_message(user_id, text=dct['text'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.text_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.text_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.text_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.text_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.text_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        await message.answer(text=dct['text'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_text_2)
    except Exception:
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.text_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_text_2), F.from_user.id.in_(ADMIN_IDS))
async def check_text_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_message(user_id, text=dct['text'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()


#Создание фото-сообщения


@router.message(F.photo, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button(message: types.Message, state: FSMContext):
    await state.update_data(photo_id=message.photo[-1].file_id)
    try:
        await state.update_data(caption=message.caption)
    except Exception:
        pass
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.photo_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.photo_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()

    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    if dct.get('caption'):
        await cb.message.answer_photo(photo=dct['photo_id'], caption=dct['caption'])
    else:
        await cb.message.answer_photo(photo=dct['photo_id'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_photo_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_photo_1), F.from_user.id.in_(ADMIN_IDS))
async def check_photo_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_photo(user_id, photo=dct['photo_id'], caption=dct['caption'])
            else:
                await bot.send_photo(user_id, photo=dct['photo_id'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.photo_add_button), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.photo_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.photo_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.photo_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.photo_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        if dct.get('caption'):
            await message.answer_photo(photo=dct['photo_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        else:
            await message.answer_photo(photo=dct['photo_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_photo_2)
    except Exception as e:
        print(e)
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.photo_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_photo_2), F.from_user.id.in_(ADMIN_IDS))
async def check_photo_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                    await bot.send_photo(user_id, photo=dct['photo_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            else:
                await bot.send_photo(user_id, photo=dct['photo_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2,
            FSMFillForm.check_photo_1, FSMFillForm.check_photo_2), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()


#Создание видео-сообщения


@router.message(F.video, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button(message: types.Message, state: FSMContext):
    await state.update_data(video_id=message.video.file_id)
    try:
        await state.update_data(caption=message.caption)
    except Exception:
        pass
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.video_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.video_add_button), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    if dct.get('caption'):
        await cb.message.answer_video(video=dct['video_id'], caption=dct['caption'])
    else:
        await cb.message.answer_video(video=dct['video_id'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_video_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_1), F.from_user.id.in_(ADMIN_IDS))
async def check_video_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_video(user_id, video=dct['video_id'], caption=dct['caption'])
            else:
                await bot.send_video(user_id, video=dct['video_id'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.video_add_button), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.video_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.video_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.video_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.video_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        if dct.get('caption'):
            await message.answer_video(video=dct['video_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        else:
            await message.answer_video(video=dct['video_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_video_2)
    except Exception as e:
        print(e)
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.video_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_2), F.from_user.id.in_(ADMIN_IDS))
async def check_video_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_video(user_id, video=dct['video_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            else:
                await bot.send_video(user_id, photo=dct['video_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


#Создание видео-кружка


@router.message(F.video_note, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def video_note_check(message: types.Message, state: FSMContext):
    await state.update_data(video_note_id=message.video_note.file_id)
    await message.answer(text='Проверьте вашу запись в кружке для отправки')
    await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_video_note_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_note_1), F.from_user.id.in_(ADMIN_IDS))
async def check_video_note_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_video_note(user_id, video_note=dct['video_note_id'])
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


# Выход из рассылки без отправки
@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2,
                       FSMFillForm.check_photo_1, FSMFillForm.check_photo_2, FSMFillForm.check_video_1,
                       FSMFillForm.check_video_2, FSMFillForm.check_video_note_1), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()


#Рассылка текста одному юзеру
@router.message(F.text == 'Sendid', StateFilter(default_state), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_1(message: types.Message, state: FSMContext):
    await message.answer(text='Введите id юзера')
    await state.set_state(FSMFillForm.send_id)


@router.message(F.text, StateFilter(FSMFillForm.send_id), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_2(message: types.Message, state: FSMContext):
    try:
        await state.update_data(user_id=int(message.text))
        await message.answer(text='Введите сообщение для отправки юзеру по id')
        await state.set_state(FSMFillForm.send_to_one)
    except Exception:
        await message.answer(text='Что-то пошло не так, проверьте корректность id. Попробуйте снова')
        await state.set_state(default_state)


@router.message(F.text, StateFilter(FSMFillForm.send_to_one), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_3(message: types.Message, state: FSMContext):
    try:
        dct = await state.get_data()
        await bot.send_message(chat_id=dct['user_id'], text=message.text)
        await message.answer(text=f"Сообщение юзеру с id {dct['user_id']} отправлено")
    except Exception:
        await message.answer(text='Что-то пошло не так, проверьте корректность id или блокировку бота юзером. Попробуйте снова')
    await state.set_state(default_state)
    await state.clear()


@router.message(F.text == 'Export', StateFilter(default_state), F.from_user.id.in_(ADMIN_IDS))
async def export_data(message: types.Message):
    db = SessionLocal()
    try:
        # Получаем все записи из базы данных
        requests = db.query(SubscriptionRequest).all()

        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Subscription Requests"

        # Заголовки столбцов
        headers = [
            "ID", "User ID", "Username", "First Name", "Last Name",
            "Channel ID", "Channel Name",
            "Request Time", "User Blocked"
        ]

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Записываем данные
        for row_num, request in enumerate(requests, 2):
            ws.cell(row=row_num, column=1, value=request.id)
            ws.cell(row=row_num, column=2, value=request.user_id)
            ws.cell(row=row_num, column=3, value=request.username or "")
            ws.cell(row=row_num, column=4, value=request.first_name or "")
            ws.cell(row=row_num, column=5, value=request.last_name or "")
            ws.cell(row=row_num, column=6, value=request.channel_id)
            ws.cell(row=row_num, column=7, value=request.channel_name or "")
            # ws.cell(row=row_num, column=8, value=request.channel_link or "")
            ws.cell(row=row_num, column=8, value=request.time_request)
            # ws.cell(row=row_num, column=10, value=request.time_confirm or "")
            ws.cell(row=row_num, column=9, value="Да" if request.user_is_block else "Нет")

        # Автоподбор ширины столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Получаем букву столбца
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Сохраняем файл в памяти
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Отправляем файл
        await message.answer_document(
            document=types.BufferedInputFile(
                excel_file.getvalue(),
                filename="subscription_requests.xlsx"
            ),
            caption="Экспорт данных о подписках"
        )

    except Exception as e:
        await message.answer(f"Произошла ошибка при экспорте: {str(e)}")
    finally:
        db.close()